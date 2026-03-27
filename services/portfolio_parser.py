"""
Portfolio Excel/CSV parser.

Supports Brewin Dolphin (Investec W&I), Hargreaves Lansdown,
Quilter, St James's Place, and generic ISIN/value formats.

Public API:
    parse_portfolio_file(file_bytes, filename) -> ParseResult
"""

import io
import re
import logging
from dataclasses import dataclass, field
from typing import Optional

logger = logging.getLogger(__name__)

# ISIN pattern: 2-letter country code + 10 alphanumeric chars
_ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")
# SEDOL pattern: 6–7 alphanumeric chars (excluding I, O)
_SEDOL_RE = re.compile(r"^[BCDFGHJKLMNPQRSTVWXYZ0-9]{6,7}$")

# UK pence threshold: if unit price > 400 for a GBP security it's likely pence
_PENCE_THRESHOLD = 400.0


@dataclass
class Holding:
    name: str
    isin: Optional[str] = None
    sedol: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None       # in GBP (pence normalised)
    value_gbp: Optional[float] = None
    cost_gbp: Optional[float] = None
    unrealised_gl: Optional[float] = None    # GBP
    unrealised_gl_pct: Optional[float] = None
    currency: str = "GBP"
    asset_class: Optional[str] = None       # Equities / Fixed Income / Funds / Cash
    resolved_ticker: Optional[str] = None
    weight_pct: Optional[float] = None


@dataclass
class ParseResult:
    holdings: list[Holding] = field(default_factory=list)
    unresolved: list[dict] = field(default_factory=list)   # {name, isin, reason}
    total_value_gbp: float = 0.0
    total_cost_gbp: float = 0.0
    format_detected: str = "unknown"
    parse_errors: list[str] = field(default_factory=list)


# ── Format detection ──────────────────────────────────────────────────────────

def _detect_format(headers: list[str], sheet_name: str) -> str:
    h = [h.lower().strip() for h in headers]
    joined = " ".join(h)

    if any(kw in joined for kw in ["brewin", "investec", "managed portfolio"]):
        return "brewin_dolphin"
    if "hargreaves" in joined or ("stock" in h and "units" in h and "cost (£)" in h):
        return "hargreaves_lansdown"
    if any(kw in joined for kw in ["quilter", "old mutual"]):
        return "quilter"
    if "st. james" in joined or "st james" in joined:
        return "st_james_place"
    # Generic: has ISIN column and value column
    if any("isin" in x for x in h) and any("value" in x for x in h):
        return "generic_isin"
    return "unknown"


# ── Column-index mapping ──────────────────────────────────────────────────────

_COL_SYNONYMS = {
    "name":       ["security", "description", "stock", "holding", "asset", "instrument", "name"],
    "isin":       ["isin"],
    "sedol":      ["sedol"],
    "quantity":   ["quantity", "units", "shares", "nominal", "qty"],
    "unit_price": ["price", "unit price", "unit cost", "price (p)", "price (£)", "bid price"],
    "value":      ["value", "market value", "current value", "value (£)", "val (£)"],
    "cost":       ["cost", "book cost", "cost (£)", "purchase cost", "total cost", "original cost"],
    "gl_abs":     ["gain/loss", "unrealised gain", "unrealised g/l", "p&l", "gain", "profit/loss"],
    "gl_pct":     ["gain/loss %", "return %", "% gain", "unrealised %"],
    "currency":   ["currency", "ccy"],
    "weight":     ["% portfolio", "weight", "%"],
}


def _map_columns(headers: list[str]) -> dict[str, int]:
    mapping = {}
    for i, raw in enumerate(headers):
        h = raw.lower().strip()
        for field_name, synonyms in _COL_SYNONYMS.items():
            if field_name not in mapping and any(syn in h for syn in synonyms):
                mapping[field_name] = i
    return mapping


# ── Cell value helpers ────────────────────────────────────────────────────────

def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("£", "").replace("%", "").replace("p", "").strip()
    if not s or s in ("-", "N/A", "n/a", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalise_pence(price: float, currency: str, name: str) -> float:
    """Convert prices quoted in pence to GBP for UK securities."""
    if currency in ("GBp", "GBX", "GBp") or (
        currency in ("GBP", "") and price > _PENCE_THRESHOLD
    ):
        return price / 100.0
    return price


# ── Row classification ────────────────────────────────────────────────────────

_SECTION_KEYWORDS = {
    "UK Equities": ["uk equit", "uk shares", "united kingdom equit"],
    "International Equities": ["international equit", "overseas equit", "global equit", "us equit"],
    "Fixed Income": ["fixed income", "bond", "gilt", "fixed interest"],
    "Funds": ["fund", "unit trust", "oeic", "investment trust", "etf"],
    "Cash": ["cash", "money market"],
    "Alternatives": ["alternative", "hedge", "private equity", "property"],
}


def _classify_row(row_vals: list, col_map: dict) -> Optional[str]:
    """Return asset class if this row is a section header, else None."""
    name_col = col_map.get("name")
    if name_col is None:
        return None
    name = _to_str(row_vals[name_col]) if name_col < len(row_vals) else None
    if not name:
        return None
    nl = name.lower()
    for asset_class, kws in _SECTION_KEYWORDS.items():
        if any(kw in nl for kw in kws):
            non_null = sum(1 for v in row_vals if v is not None and str(v).strip())
            if non_null <= 2:  # section header rows have very few populated cells
                return asset_class
    return None


def _is_total_row(row_vals: list, col_map: dict) -> bool:
    """Skip total/subtotal/summary rows."""
    name_col = col_map.get("name")
    if name_col is None:
        return False
    name = _to_str(row_vals[name_col]) if name_col < len(row_vals) else None
    if not name:
        return False
    nl = name.lower()
    return any(kw in nl for kw in ["total", "subtotal", "grand total", "net asset", "portfolio total"])


def _is_data_row(row_vals: list, col_map: dict) -> bool:
    """Check if row looks like a real holding (has name + at least one numeric)."""
    name_col = col_map.get("name")
    if name_col is None or name_col >= len(row_vals):
        return False
    name = _to_str(row_vals[name_col])
    if not name or len(name) < 3:
        return False
    # Must have at least one numeric in value/quantity/cost columns
    for field_name in ("value", "quantity", "cost"):
        idx = col_map.get(field_name)
        if idx is not None and idx < len(row_vals):
            if _to_float(row_vals[idx]) is not None:
                return True
    return False


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_sheet_rows(rows: list[list], filename: str) -> ParseResult:
    result = ParseResult()

    # Find the header row (first row with 3+ non-empty string cells that look like labels)
    header_row_idx = None
    for i, row in enumerate(rows[:20]):
        non_empty = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(non_empty) >= 3 and not any(_ISIN_RE.match(v) for v in non_empty):
            header_row_idx = i
            break

    if header_row_idx is None:
        result.parse_errors.append("Could not find header row in sheet")
        return result

    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_idx]]
    result.format_detected = _detect_format(headers, filename)
    col_map = _map_columns(headers)

    if "name" not in col_map:
        result.parse_errors.append("Could not locate security name column")
        return result

    current_asset_class = None

    for row in rows[header_row_idx + 1:]:
        # Pad/trim row to header length
        row_vals = list(row) + [None] * max(0, len(headers) - len(row))

        # Skip empty rows
        non_empty = [v for v in row_vals if v is not None and str(v).strip()]
        if not non_empty:
            continue

        # Detect section headers
        section = _classify_row(row_vals, col_map)
        if section:
            current_asset_class = section
            continue

        # Skip total/summary rows
        if _is_total_row(row_vals, col_map):
            continue

        # Skip rows that don't look like holdings
        if not _is_data_row(row_vals, col_map):
            continue

        def _get(field_name):
            idx = col_map.get(field_name)
            return row_vals[idx] if idx is not None and idx < len(row_vals) else None

        name = _to_str(_get("name"))
        if not name:
            continue

        isin_raw = _to_str(_get("isin"))
        sedol_raw = _to_str(_get("sedol"))
        isin = isin_raw if isin_raw and _ISIN_RE.match(isin_raw.upper()) else None
        sedol = sedol_raw if sedol_raw and _SEDOL_RE.match(sedol_raw.upper().replace(" ", "")) else None

        currency = _to_str(_get("currency")) or "GBP"
        quantity = _to_float(_get("quantity"))
        unit_price_raw = _to_float(_get("unit_price"))
        value_raw = _to_float(_get("value"))
        cost_raw = _to_float(_get("cost"))
        gl_abs = _to_float(_get("gl_abs"))
        gl_pct = _to_float(_get("gl_pct"))
        weight_pct = _to_float(_get("weight"))

        # Normalise pence → GBP
        unit_price = _normalise_pence(unit_price_raw, currency, name) if unit_price_raw is not None else None

        # Infer value from quantity × price if missing
        if value_raw is None and unit_price is not None and quantity is not None:
            value_raw = unit_price * quantity

        holding = Holding(
            name=name,
            isin=isin.upper() if isin else None,
            sedol=sedol.upper() if sedol else None,
            quantity=quantity,
            unit_price=unit_price,
            value_gbp=value_raw,
            cost_gbp=cost_raw,
            unrealised_gl=gl_abs,
            unrealised_gl_pct=gl_pct,
            currency=currency,
            asset_class=current_asset_class,
            weight_pct=weight_pct,
        )
        result.holdings.append(holding)

    # Compute totals
    result.total_value_gbp = sum(h.value_gbp for h in result.holdings if h.value_gbp)
    result.total_cost_gbp = sum(h.cost_gbp for h in result.holdings if h.cost_gbp)

    # Compute weights if not present
    if result.total_value_gbp > 0:
        for h in result.holdings:
            if h.weight_pct is None and h.value_gbp:
                h.weight_pct = (h.value_gbp / result.total_value_gbp) * 100

    return result


# ── Excel & CSV entry points ───────────────────────────────────────────────────

def _parse_excel(file_bytes: bytes, filename: str) -> ParseResult:
    try:
        import openpyxl
    except ImportError:
        r = ParseResult()
        r.parse_errors.append("openpyxl not installed")
        return r

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Prefer sheets with "portfolio", "valuation", "holdings" in the name
    preferred_order = []
    for name in wb.sheetnames:
        nl = name.lower()
        if any(kw in nl for kw in ["portfolio", "valuation", "holdings", "assets"]):
            preferred_order.insert(0, name)
        elif any(kw in nl for kw in ["transaction", "income", "allocation"]):
            pass  # skip secondary sheets
        else:
            preferred_order.append(name)

    best_result = ParseResult()
    best_result.parse_errors.append("No suitable sheet found in workbook")

    for sheet_name in preferred_order:
        ws = wb[sheet_name]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        result = _parse_sheet_rows(rows, sheet_name)
        if len(result.holdings) > len(best_result.holdings):
            best_result = result
            if len(result.holdings) >= 3:
                break  # Good enough

    return best_result


def _parse_csv(file_bytes: bytes, filename: str) -> ParseResult:
    import csv
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    return _parse_sheet_rows([[v for v in row] for row in rows], filename)


# ── ISIN/SEDOL resolution ────────────────────────────────────────────────────

def _resolve_tickers(result: ParseResult) -> None:
    """
    Resolve ISINs/SEDOLs to tickers via OpenFIGI.
    Mutates result in-place.
    """
    from services.openfigi_service import resolve_identifier

    for holding in result.holdings:
        identifier = holding.isin or holding.sedol
        if not identifier:
            result.unresolved.append({
                "name": holding.name,
                "reason": "no ISIN or SEDOL",
            })
            continue

        id_type = "ID_ISIN" if holding.isin else "ID_SEDOL"
        ticker = resolve_identifier(identifier, id_type)
        if ticker:
            holding.resolved_ticker = ticker
        else:
            result.unresolved.append({
                "name": holding.name,
                "isin": holding.isin,
                "sedol": holding.sedol,
                "reason": "OpenFIGI lookup failed",
            })


# ── Public API ────────────────────────────────────────────────────────────────

def parse_portfolio_file(file_bytes: bytes, filename: str, resolve: bool = True) -> ParseResult:
    """
    Parse an Excel (.xlsx) or CSV portfolio export.

    Args:
        file_bytes: Raw file content
        filename:   Original filename (used for format detection)
        resolve:    Whether to call OpenFIGI to resolve tickers (default True)

    Returns:
        ParseResult with holdings list and metadata
    """
    fn_lower = filename.lower()
    if fn_lower.endswith(".csv"):
        result = _parse_csv(file_bytes, filename)
    elif fn_lower.endswith((".xlsx", ".xls", ".xlsm")):
        result = _parse_excel(file_bytes, filename)
    else:
        result = ParseResult()
        result.parse_errors.append(f"Unsupported file type: {filename}")
        return result

    logger.info(
        "Parsed %d holdings from %s (format: %s, total: £%.0f)",
        len(result.holdings), filename, result.format_detected, result.total_value_gbp
    )

    if resolve and result.holdings:
        _resolve_tickers(result)

    resolved_count = sum(1 for h in result.holdings if h.resolved_ticker)
    logger.info(
        "Ticker resolution: %d/%d resolved, %d unresolved",
        resolved_count, len(result.holdings), len(result.unresolved)
    )

    return result
